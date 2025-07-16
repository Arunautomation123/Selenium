from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import time
import openpyxl
from openpyxl.styles import PatternFill

options = webdriver.ChromeOptions()
options.add_experimental_option(
    "prefs",
    {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.default_content_setting_values.notifications": 2
        # with 2 should disable notifications
    },
)
options.add_argument('--disable-notifications')
options.add_argument("--disable-infobars")
options.add_argument("--disable-extensions")
options.add_argument("start-maximized")

service_obj = Service(executable_path=r'D:\Projects\driver\chromedriver.exe')

driver = webdriver.Chrome(service=service_obj, options=options)

driver.implicitly_wait(10)

driver.get('https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html')

driver.maximize_window()

path = r'D:\Projects\documents\Calc_data.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
color = PatternFill(start_color='00ff00', end_color='00ff00', fill_type='solid')
red_color = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')

for i in range(2, sheet.max_row+1):
    princ = sheet.cell(i, 1).value
    roi = sheet.cell(i, 2).value
    per1 = sheet.cell(i, 3).value
    per2 = sheet.cell(i, 4).value
    frq = sheet.cell(i, 7).value
    exp_val = sheet.cell(i, 6).value
    mat_amt = sheet.cell(i, 5).value

    driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(princ)
    driver.find_element(By.XPATH, "//input[@id='interest']").send_keys(roi)
    driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(per1)
    sel = driver.find_element(By.XPATH, "//select[@id='tenurePeriod']")
    sel2 = driver.find_element(By.XPATH, "//select[@id='frequency']")
    element = Select(sel)
    element.select_by_visible_text(per2)

    element2 = Select(sel2)
    element2.select_by_visible_text(frq)

    time.sleep(3)

    driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]/img").click()

    app_cal_val = driver.find_element(By.XPATH, "//span[@class='gL_27']/strong").text

    print(f"Mautiry amount: {app_cal_val}")

    if float(app_cal_val) == float(mat_amt):
        print("Calculated amount matched with value from excel")
        sheet.cell(i, 8).value = 'Pass'
        sheet.cell(i, 8).fill = color
    else:
        print("Calculated amount did not match with value from excel")
        sheet.cell(i, 8).value = 'Fail'
        sheet.cell(i, 8).fill = red_color

    driver.find_element(By.XPATH, "//img[@class='PL5']").click()
    time.sleep(3)

    workbook.save(path)

time.sleep(5)

driver.quit()