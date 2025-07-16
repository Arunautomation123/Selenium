from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import time
import openpyxl
from openpyxl.styles import PatternFill

# service_obj = Service(executable_path=r'D:\Projects\driver\chromedriver.exe')
# driver = webdriver.Chrome(service=service_obj)

# when repeated data needs to be tested - data driven testing

# 1. Read data from excel
# file ---- workbook -- sheet -- row -- cells
# path = r'D:\Projects\documents\data.xlsx'
# workbook = openpyxl.load_workbook(path)
# sheet = workbook['Sheet1']
#
# rows = sheet.max_row
# columns = sheet.max_column
#
# print(f"No of rows: {rows}, No of columns: {columns}")
#
# for r in range(1, rows+1):
#     for c in range(1, columns+1):
#         print(f"{sheet.cell(r, c).value}", end='  ')
#     print()


## 2.write same data into excel
path = r'D:\Projects\documents\test_data.xlsx'
workbook = openpyxl.load_workbook(path)
# if only one sheet is present - sheet = workbook.active
sheet = workbook['Sheet1']

# fill RGB values
greenfil = PatternFill(start_color='00ff00', fill_type='solid')

for i in range(1, 5):
    for j in range(1, 4):
        sheet.cell(i, j).value = i*j
        if i > 3:
            sheet.cell(i, j).fill=greenfil

workbook.save(path)

## 3.write unique data into excel
# dict1 = {'name':'arun', 'age': 27, 'place':'bellary'}
# dict2 = {'name':'shiva', 'age': 26, 'place':'bangalore'}
# dict3 = {'name':'sanju', 'age': 28, 'place':'hubli'}
# path = r'D:\Projects\documents\test_data1.xlsx'
# workbook = openpyxl.load_workbook(path)
# # if only one sheet is present - sheet = workbook.active
# sheet = workbook['Sheet1']
#
# data = [["Name", "Age", "PLcae"],
#         ['Arun', 27, 'bellary'],
#         ['sanju', 27, 'bellary']]
#
# for r in data:
#     sheet.append(r)
#
#
#
# workbook.save(path)