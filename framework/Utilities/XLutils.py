import openpyxl

def getRowcnt(file, sheet):
    book = openpyxl.load_workbook(file)
    sheetname = book[sheet]
    return sheetname.max_row

def getColcnt(file, sheet):
    book = openpyxl.load_workbook(file)
    sheetname = book[sheet]
    return sheetname.max_column

def read_data(file, sheet, row, column):
    book = openpyxl.load_workbook(file)
    sheetname = book[sheet]
    return sheetname.cell(row, column).value

def write_data(file, sheet, row, column, data):
    book = openpyxl.load_workbook(file)
    sheetname = book[sheet]
    sheetname.cell(row, column).value = data
    book.save(file)